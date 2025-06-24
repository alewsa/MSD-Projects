''' 
COMP 467 Project 4 The Crucible
Alyssa Andres

Goal: 
1. Revisit Proj 1
2. Add argparse to input baselight file (--baselight),  xytech (--xytech) from proj1
3. Populate new database with 2 collections: One for Baselight (Folder/Frames) and Xytech (Workorder/Location) 
4. Download Demo reel: https://mycsun.box.com/s/wy1sit1vne1gmg8l1psbbyetp73ls4hf
5. Run script with new argparse command --process <video file>  
6. From (5) Call the populated database from (3), find all ranges only that fall in the length of video from (4)
7. Using ffmpeg or 3rd party tool of your choice, to extract timecode from video and write your own timecode method to convert marks to timecode
8. New argparse--output parameter for XLS with flag from (5) should export same CSV export as proj 1 (matching xytech/baselight locations), 
   but in XLS with new column from files found from (6) and export their timecode ranges as well
9. Create Thumbnail (96x74) from each entry in (6), but middle most frame or closest to. Add to XLS file to it's corresponding range in new column 
10. Render out each shot from (6) using (7) and upload them using API to Vimeo (see slide 11)
11. Create CSV file and show all ranges/individual frames that were not uploaded from  (10) (so show location and frame/ranges)
 
Deliverables:
Copy/Paste code
Excel file with new columns noted on Solve (8) and (9)
Screenshot of Vimeo account (10)
CSV export of unused frames (11)

Type in Terminal for Testing:
python Andres_Project4.py --baselight Baselight_export_spring2025.txt --xytech Xytech_spring2025.txt --process FROM-ZERO_Demo-REEL_V01-04.mp4 --output test_report.xlsx

'''

import argparse
import csv
import os
import time
import pandas as pd
from pymongo import MongoClient
import ffmpeg
from PIL import Image
import vimeo
from openpyxl import Workbook #used to create Excel file
from openpyxl.drawing.image import Image as XLImage #to see thumbnails in Excel file
from openpyxl.utils import get_column_letter #converts column index (1,2,3) to Excel letters (A,B,C) for setting column width

#Vimeo credentials (access token has upload and edit scope)
CLIENT_ID     = 'f90cb69355ff88b5c992d7ca064c8a43fdba759a'
CLIENT_SECRET = 'gaNCEixtkGH5MKxBd/gtOZpi9muFMMnGQJJ+mXp4R2FjmlWn2pvtCICN18t/yKKNVLvXwtMHB7PmvV5D6fCNSyageFdNcJwsy4LQzDzVez/g9KwVe9MP+AIqcvyrJl9k'
ACCESS_TOKEN  = '025ea95c21dcfa9f2846dd6a4cfe4176' 

#Connect to MongoDB
client = MongoClient('mongodb://localhost:27017/')
db = client['proj4_db']

#Convert frame number to timecode string (HH:MM:SS:FF)
def frame_to_timecode(frame, fps=24):
    hours = frame // (3600 * fps)
    minutes = (frame // (60 * fps)) % 60
    seconds = (frame // fps) % 60
    frames = frame % fps
    return f"{hours:02}:{minutes:02}:{seconds:02}:{frames:02}"

#Group consecutive numbers into sublists
def split_into_frame_groups(frames):
    if not frames:
        return []
    frames = sorted(set(frames))
    groups = []
    current = [frames[0]]
    for f in frames[1:]:
        if f == current[-1] + 1:
            current.append(f)
        else:
            groups.append(current)
            current = [f]
    groups.append(current)
    return groups

#Match Baselight path to closest Xytech location
def map_to_xytech_location(baselight_path, xytech_locations):
    for xyt in xytech_locations:
        if baselight_path.split("/production/")[-1] in xyt:
            return xyt
        if baselight_path.split("/dogman/")[-1] in xyt:
            return xyt
    return baselight_path

#Sort unused frames from smallest frame(s) to biggest frame(s)
def sort_by_first_frame(entry):
    if entry['frames']:
        return entry['frames'][0]
    else:
        return float('inf')

#Upload clips to Vimeo with title and description. Retry upload if Vimeo API limit is reached.
def upload_to_vimeo(filepath, title, description, client_id, client_secret, access_token):
    if not os.path.isfile(filepath):
        print(f"[ERROR] Clip file not found: {filepath}") 
        return None

    v = vimeo.VimeoClient(token=access_token, key=client_id, secret=client_secret)

    for attempt in range(3): #attempt to upload clip 3 times after hitting API rate limit
        try:
            uri = v.upload(filepath)
            v.patch(uri, data={'name': title, 'description': description})
            print(f"[SUCCESS] Uploaded to Vimeo: https://vimeo.com{uri}")
            time.sleep(3)  #Pause after each successful upload
            return f"https://vimeo.com{uri}"
        except Exception as e: #wait for 60 secs and retry uploading
            print(f"[RETRY {attempt + 1}] Upload failed for {filepath}: {e}")
            if "too many api requests" in str(e).lower():
                print("[WAIT] Vimeo API rate limit hit. Wait for 60 seconds to try uploading again...")
                time.sleep(60)
            else:
                time.sleep(3)

    #print if upload fails after third attempt
    print(f"[ERROR] Failed to upload after 3 attempts: {filepath}")
    return None

#Argparse functions
def main():
    parser = argparse.ArgumentParser()
    parser.add_argument('--baselight', required=True)
    parser.add_argument('--xytech', required=True)
    parser.add_argument('--process', required=True)
    parser.add_argument('--output', required=True)
    args = parser.parse_args()

    #Input file checks (print statements for baselight and xytech files if found or not)
    if not os.path.isfile(args.baselight):
        print(f"[ERROR] Baselight file not found: {args.baselight}")
        return
    else:
        print(f"[INFO] Baselight file loaded: {args.baselight}")

    if not os.path.isfile(args.xytech):
        print(f"[ERROR] Xytech file not found: {args.xytech}")
        return
    else:
        print(f"[INFO] Xytech file loaded: {args.xytech}")

    #Output folders (for storing clips and thumbnails)
    output_dir = "output"
    thumbnail_dir = os.path.join(output_dir, "thumbnails")
    clip_dir = os.path.join(output_dir, "clips")
    os.makedirs(thumbnail_dir, exist_ok=True)
    os.makedirs(clip_dir, exist_ok=True)

    #Reset database
    baselight_col = db['baselight']
    xytech_col = db['xytech']
    baselight_col.delete_many({})
    xytech_col.delete_many({})

    #Load Xytech locations
    xytech_locations = []
    with open(args.xytech) as xy_file:
        for line in xy_file:
            line = line.strip()
            if line.startswith('/'):
                xytech_locations.append(line)
                xytech_col.insert_one({'location': line})

    #Load Baselight data
    with open(args.baselight) as bl_file:
        for line in bl_file:
            line = line.strip()
            if not line:
                continue
            parts = line.split()
            if len(parts) < 2:
                continue
            path, frames = parts[0], list(map(int, parts[1:]))
            baselight_col.insert_one({'path': path, 'frames': frames})

    #Get video info (video duration and total frames)
    probe = ffmpeg.probe(args.process)
    duration = float(probe['format']['duration'])
    video_fps = 24
    total_frames = int(duration * video_fps)

    #Lists to keep clips being used and frames being skipped
    report_data = []
    unused_frames = []

    #Go through all the Baselight entries and observe which frame groups are valid
    for entry in baselight_col.find({}):
        all_frames = entry['frames']
        if not all_frames:
            continue

        frame_groups = split_into_frame_groups(all_frames)
        used_frames = []

        for group in frame_groups:
            start_frame = group[0]
            end_frame = group[-1]

            if len(group) == 1:
                continue

            if end_frame <= total_frames:
                used_frames.extend(group)

                #Create thumbnail
                thumbnail_frame = (start_frame + end_frame) // 2
                thumbnail_time = thumbnail_frame / video_fps
                thumbnail_path = os.path.join(thumbnail_dir, f"thumbnail_{thumbnail_frame}.jpg")
                (
                    ffmpeg.input(args.process, ss=thumbnail_time)
                    .output(thumbnail_path, vframes=1)
                    .overwrite_output()
                    .run(quiet=True)
                )
                img = Image.open(thumbnail_path)
                img.thumbnail((96, 74))
                img.save(thumbnail_path)

                #Add to Excel file
                timecode_range = f"{frame_to_timecode(start_frame)} - {frame_to_timecode(end_frame)}"
                report_data.append({
                    "Location": map_to_xytech_location(entry['path'], xytech_locations),
                    "Frames": f"{start_frame}-{end_frame}",
                    "Timecode": timecode_range,
                    "Thumbnail": thumbnail_path
                })

                #Create and upload clip
                clip_path = os.path.join(clip_dir, f"clip_{start_frame}_{end_frame}.mp4")
                start_sec = start_frame / video_fps
                duration_sec = (end_frame - start_frame + 1) / video_fps

                #Pad short clips using tpad to reach at least 1 second (to avoid upload failure in Vimeo)
                (
                    ffmpeg.input(args.process, ss=start_sec, t=duration_sec)
                    .filter('tpad', stop_mode='clone', stop_duration=max(1.0 - duration_sec, 0))
                    .output(clip_path, format='mp4', movflags='faststart')
                    .overwrite_output()
                    .run(quiet=False)
                )

                #Wait and make sure video is processed
                for _ in range(6):
                    if os.path.exists(clip_path) and os.path.getsize(clip_path) > 0:
                        break
                    time.sleep(0.5)

                #Check if file for clips exist
                try:
                    with open(clip_path, 'rb') as f:
                        f.read()
                except Exception as e:
                    print(f"[ERROR] Could not verify clip file: {e}")
                    continue

                time.sleep(1.0)

                #Upload clips to Vimeo
                upload_to_vimeo(
                    clip_path,
                    title=f"Clip {start_frame}-{end_frame}",
                    description="Clip upload for COMP 467 - Project 4 The Crucible",
                    client_id=CLIENT_ID,
                    client_secret=CLIENT_SECRET,
                    access_token=ACCESS_TOKEN
                )
            else: #unused frames from the video
                unused_frames.append({
                    "path": map_to_xytech_location(entry['path'], xytech_locations),
                    "frames": group
                })

        #Handle unused single frames
        singles = [
            f for f in all_frames
            if f not in used_frames
            and all(f not in grp for grp in frame_groups if len(grp) > 1)
        ]
        if singles:
            unused_frames.append({
                "path": map_to_xytech_location(entry['path'], xytech_locations),
                "frames": singles
            })

    #Create Excel file
    wb = Workbook()
    ws = wb.active
    ws.title = "Frame Report"
    headers = ["Location", "Frames", "Timecode", "Thumbnail"]
    ws.append(headers)

    column_widths = [50, 20, 25, 15] #avoids data from getting cramped up together
    for i, width in enumerate(column_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width

    for idx, entry in enumerate(report_data, start=2):
        ws.cell(row=idx, column=1, value=entry["Location"])
        ws.cell(row=idx, column=2, value=entry["Frames"])
        ws.cell(row=idx, column=3, value=entry["Timecode"])
        if os.path.exists(entry["Thumbnail"]):
            xl_img = XLImage(entry["Thumbnail"])
            xl_img.width = 96
            xl_img.height = 74
            ws.row_dimensions[idx].height = 60
            ws.add_image(xl_img, f"D{idx}")

    try:
        wb.save(args.output)
        print(f"[SUCCESS] Excel file saved as {args.output}")
    except Exception as e:
        print(f"[ERROR] Failed to save Excel: {e}")

    #Create CSV file for unused frames
    unused_frames.sort(key=sort_by_first_frame)
    try:
        with open("unused_frames.csv", "w", newline='') as csvfile:
            writer = csv.writer(csvfile)
            writer.writerow(["Location", "Unused Frames"])
            for entry in unused_frames:
                writer.writerow([entry['path'], ','.join(map(str, entry['frames']))])
        print("[SUCCESS] CSV with unused frames saved as unused_frames.csv")
    except Exception as e:
        print(f"[ERROR] Failed to save CSV: {e}")

if __name__ == "__main__":
    main()